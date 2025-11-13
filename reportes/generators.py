# reportes/generators.py - CORREGIDO
import datetime
from decimal import Decimal
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

def _limpiar_valor(valor):
    """Convierte valores especiales (Decimal, Fecha, None) a strings legibles."""
    if isinstance(valor, Decimal):
        return f"Bs. {valor:,.2f}"
    if isinstance(valor, (datetime.date, datetime.datetime)):
        return valor.strftime("%Y-%m-%d")
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "Sí" if valor else "No"
    return str(valor)

def _formatear_encabezado(header):
    """Formatea los encabezados para mejor presentación"""
    mapeo = {
        'id': 'ID',
        'precio_contado': 'Precio Contado',
        'precio_cuota': 'Precio Cuota', 
        'fecha_registro': 'Fecha Registro',
        'is_active': 'Activo',
        'garantia_meses': 'Garantía (Meses)',
        'subcategoria__nombre': 'Subcategoría',
        'marca__nombre': 'Marca',
        'categoria__nombre': 'Categoría',
        'usuario__username': 'Usuario',
        'total': 'Total',
        'cantidad': 'Cantidad',
        'precio_unitario': 'Precio Unitario',
        'subtotal': 'Subtotal',
        'estado': 'Estado',
        'fecha_pago': 'Fecha Pago',
        'monto': 'Monto',
        'numero_cuota': 'N° Cuota',
        'fecha_vencimiento': 'Fecha Vencimiento',
    }
    header_limpio = header.replace('_', ' ').title()
    return mapeo.get(header, header_limpio)

# ===================================================================
# --- GENERADOR DE REPORTE EXCEL (OPENPYXL) - CORREGIDO ---
# ===================================================================
def generar_reporte_excel(data, interpretacion):
    """
    Genera un archivo Excel (xlsx) en memoria para ecommerce
    """
    prompt_titulo = interpretacion.get('prompt', 'Reporte Ecommerce')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_ecommerce_{datetime.date.today()}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    if not data:
        ws['A1'] = "No se encontraron datos para este reporte."
        wb.save(response)
        return response

    # --- Estilos CORREGIDOS ---
    font_titulo = Font(bold=True, size=14)
    font_header = Font(bold=True, color="FFFFFF")
    
    # ✅ CORRECCIÓN: Usar PatternFill con color HEX correctamente
    fill_header = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # --- Título ---
    num_columns = len(data[0].keys()) if data else 1
    ws['A1'] = prompt_titulo
    ws.merge_cells(f'A1:{get_column_letter(num_columns)}1')
    ws['A1'].font = font_titulo
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Fecha de generación
    ws['A2'] = f"Generado el: {datetime.date.today()}"
    ws.merge_cells(f'A2:{get_column_letter(num_columns)}2')
    ws['A2'].alignment = Alignment(horizontal='center')

    # --- Encabezados ---
    headers = list(data[0].keys())
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=_formatear_encabezado(header_title))
        cell.font = font_header
        cell.fill = fill_header  # ✅ Usar el PatternFill corregido
        cell.alignment = alignment_center
        cell.border = border_thin

    # --- Datos ---
    row_num = 5
    for row_data in data:
        for col_num, header in enumerate(headers, 1):
            valor = row_data.get(header)
            cell = ws.cell(row=row_num, column=col_num, value=_limpiar_valor(valor))
            cell.border = border_thin
            
            # Alineación especial para números
            if isinstance(valor, (int, float, Decimal)):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
        row_num += 1

    # --- Ajustar columnas ---
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        column = ws[column_letter]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response

# ===================================================================
# --- GENERADOR DE REPORTE PDF (REPORTLAB) - CORREGIDO ---
# ===================================================================
def generar_reporte_pdf(data, interpretacion):
    """
    Genera un archivo PDF en memoria para ecommerce
    """
    prompt_titulo = interpretacion.get('prompt', 'Reporte Ecommerce')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_ecommerce_{datetime.date.today()}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), 
                          topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()

    if not data:
        story.append(Paragraph("No se encontraron datos para este reporte.", styles['Normal']))
        doc.build(story)
        return response

    # --- Título ---
    story.append(Paragraph(prompt_titulo, styles['h1']))
    story.append(Paragraph(f"Generado el: {datetime.date.today()}", styles['Normal']))
    story.append(Spacer(1, 0.25*inch))

    # --- Preparar datos de tabla ---
    headers = list(data[0].keys())
    clean_headers = [_formatear_encabezado(h) for h in headers]
    
    table_data = [clean_headers] + [
        [_limpiar_valor(row.get(header)) for header in headers] for row in data
    ]

    # --- Crear tabla ---
    t = Table(table_data, repeatRows=1)

    # --- Estilo de tabla ---
    style = TableStyle([
        # ✅ CORRECCIÓN: Usar colors.HexColor correctamente
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ])
    
    # Aplicar estilo alternado para filas
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            # ✅ CORRECCIÓN: Usar colors.whitesmoke o colors.lightgrey
            style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)
    
    t.setStyle(style)
    story.append(t)

    doc.build(story)
    return response

# En tu generators.py - Mejora la función generar_reporte_pdf

# En tu generators.py - CORREGIDO

def generar_reporte_cliente_pdf(data, interpretacion):
    """
    Genera un archivo PDF en memoria para ecommerce - CORREGIDO
    """
    prompt_titulo = interpretacion.get('prompt', 'Reporte Ecommerce')
    tipo_reporte = interpretacion.get('tipo_reporte', '')
    total_resultados = interpretacion.get('total_resultados', len(data))
    fecha_consulta = interpretacion.get('fecha_consulta', datetime.date.today())

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{datetime.date.today()}.pdf"'

    # Usar página vertical para mejor legibilidad
    doc = SimpleDocTemplate(response, pagesize=letter, 
                          topMargin=0.5*inch, bottomMargin=0.5*inch,
                          leftMargin=0.5*inch, rightMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()

    if not data:
        story.append(Paragraph("No se encontraron datos para este reporte.", styles['Normal']))
        doc.build(story)
        return response

    # --- Título y información del reporte ---
    title_style = styles['Heading1']
    title_style.alignment = 1  # Centrado
    story.append(Paragraph("Reporte de Compras", title_style))
    story.append(Spacer(1, 0.1*inch))
    
    # Información del reporte
    info_style = styles['Normal']
    info_style.alignment = 1
    story.append(Paragraph(f"<b>Consulta:</b> {prompt_titulo}", info_style))
    story.append(Paragraph(f"<b>Tipo:</b> {tipo_reporte.title() if tipo_reporte else 'General'}", info_style))
    story.append(Paragraph(f"<b>Total de registros:</b> {total_resultados}", info_style))
    story.append(Paragraph(f"<b>Fecha de generación:</b> {fecha_consulta}", info_style))
    story.append(Spacer(1, 0.2*inch))

    # --- Preparar datos de tabla ---
    headers = list(data[0].keys()) if data else []
    clean_headers = [_formatear_encabezado(h) for h in headers]
    
    table_data = [clean_headers] + [
        [_limpiar_valor(row.get(header)) for header in headers] for row in data
    ]

    # --- CORRECCIÓN: Calcular anchos de columna apropiados ---
    if headers:
        # Calcular anchos basados en el contenido
        col_widths = []
        for i, header in enumerate(headers):
            # Ancho mínimo basado en el encabezado
            header_width = len(clean_headers[i]) * 7  # Aproximadamente 7 puntos por carácter
            # Verificar el contenido más largo en esta columna
            max_content_width = header_width
            for row in data:
                content = str(_limpiar_valor(row.get(header, '')))
                content_width = len(content) * 6  # Aproximadamente 6 puntos por carácter
                if content_width > max_content_width:
                    max_content_width = content_width
            
            # Limitar el ancho máximo a 200 puntos
            col_width = min(max_content_width + 10, 200)
            col_widths.append(col_width)
    else:
        col_widths = None

    # --- Crear tabla con anchos calculados ---
    t = Table(table_data, repeatRows=1, colWidths=col_widths)

    # --- Estilo de tabla mejorado ---
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ])
    
    # Ajustar fuentes para datos
    if headers:
        for i in range(len(headers)):
            style.add('FONTSIZE', (i, 1), (i, -1), 8)
    
    t.setStyle(style)
    story.append(t)

    # --- Pie de página ---
    story.append(Spacer(1, 0.2*inch))
    footer_style = styles['Normal']
    footer_style.alignment = 1
    footer_style.fontSize = 8
    footer_style.textColor = colors.grey
    story.append(Paragraph("Sistema Ecommerce - Reporte generado automáticamente", footer_style))

    doc.build(story)
    return response